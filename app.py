"""
Smart Retail System - Products Module
Tanzania POS & Inventory Management
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import os
import json
import csv
from decimal import Decimal
from werkzeug.utils import secure_filename
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ==================== CONFIGURATION ====================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(BASE_DIR, "smart_retail.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key = 'smart-retail-tz-dev-secret'
app.config['JSON_SORT_KEYS'] = False
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'uploads')
app.config['ALLOWED_EXTENSIONS'] = {'csv', 'xlsx', 'xls'}

db = SQLAlchemy(app)


# ==================== DATABASE MODELS ====================

class Product(db.Model):
    """
    Product Model - Core data structure for retail products
    
    Designed for Tanzanian retail market:
    - Flexible SKU system (manual or auto-generated)
    - Prices in TZS (Tanzania Shillings)
    - Reorder level for inventory management
    - Soft-delete via is_active flag
    """
    __tablename__ = 'products'
    
    product_id = db.Column(db.Integer, primary_key=True)
    sku = db.Column(db.String(50), unique=True, nullable=False, index=True)
    name = db.Column(db.String(255), nullable=False, index=True)
    category = db.Column(db.String(100), nullable=False, index=True)
    cost_price_tzs = db.Column(db.Numeric(12, 2), nullable=False)
    selling_price_tzs = db.Column(db.Numeric(12, 2), nullable=False)
    quantity_in_stock = db.Column(db.Integer, default=0, nullable=False, index=True)
    reorder_level = db.Column(db.Integer, default=10, nullable=False)
    description = db.Column(db.Text)
    image_url = db.Column(db.String(500))
    is_active = db.Column(db.Boolean, default=True, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def to_dict(self):
        """Convert product to dictionary for API responses"""
        cost = float(self.cost_price_tzs)
        selling = float(self.selling_price_tzs)
        margin = ((selling - cost) / cost * 100) if cost > 0 else 0
        
        return {
            'product_id': self.product_id,
            'sku': self.sku,
            'name': self.name,
            'category': self.category,
            'cost_price_tzs': cost,
            'selling_price_tzs': selling,
            'margin_percent': round(margin, 2),
            'quantity_in_stock': self.quantity_in_stock,
            'reorder_level': self.reorder_level,
            'stock_status': 'Low Stock' if self.quantity_in_stock < self.reorder_level else 'In Stock',
            'description': self.description,
            'image_url': self.image_url,
            'is_active': self.is_active,
            'created_at': self.created_at.isoformat(),
            'updated_at': self.updated_at.isoformat()
        }
    
    def __repr__(self):
        return f'<Product {self.sku}: {self.name}>'
    
    def get_current_stock(self):
        """Calculate current stock from inventory movements (not direct edit)"""
        movements = InventoryMovement.query.filter_by(sku=self.sku).all()
        
        stock = 0
        for movement in movements:
            if movement.movement_type in ['STOCK_IN', 'RETURN', 'ADJUSTMENT']:
                stock += movement.quantity
            elif movement.movement_type in ['SALE', 'TRANSFER_OUT']:
                stock -= movement.quantity
        
        return max(0, stock)  # Never negative


class InventoryMovement(db.Model):
    """
    Inventory Movement Model - Audit trail for all stock changes
    
    Tracks all inventory transactions:
    - STOCK_IN: Goods received from supplier
    - SALE: Product sold at POS
    - ADJUSTMENT: Manual correction (damage, loss, recount)
    - RETURN: Customer return
    - TRANSFER_OUT: Stock moved to another location (future)
    
    Key Principle: Stock is NEVER edited directly - it's always the sum of movements
    """
    __tablename__ = 'inventory_movements'
    
    movement_id = db.Column(db.Integer, primary_key=True)
    sku = db.Column(db.String(50), db.ForeignKey('products.sku'), nullable=False, index=True)
    movement_type = db.Column(db.String(20), nullable=False)  # STOCK_IN, SALE, ADJUSTMENT, RETURN, TRANSFER_OUT
    quantity = db.Column(db.Integer, nullable=False)  # Always positive, sign determined by movement_type
    reference = db.Column(db.String(100))  # Supplier invoice, POS receipt, etc
    notes = db.Column(db.Text)
    created_by = db.Column(db.String(100))  # User who recorded movement (optional)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    
    # Relationship
    product = db.relationship('Product', backref='movements')
    
    def to_dict(self):
        """Convert movement to dictionary for API responses"""
        return {
            'movement_id': self.movement_id,
            'sku': self.sku,
            'product_name': self.product.name if self.product else self.sku,
            'movement_type': self.movement_type,
            'quantity': self.quantity,
            'reference': self.reference,
            'notes': self.notes,
            'created_by': self.created_by,
            'created_at': self.created_at.isoformat()
        }
    
    def __repr__(self):
        return f'<Movement {self.movement_type}: {self.sku} x{self.quantity}>'


class Sale(db.Model):
    """
    Sale Model - POS transaction header
    
    Represents a complete sale transaction:
    - Header with receipt number, cashier, payment details
    - Links to multiple sale_items (lines)
    - Tracks payment status and method
    - Audit-safe: voided sales marked with status, never hard-deleted
    """
    __tablename__ = 'sales'
    
    sale_id = db.Column(db.Integer, primary_key=True)
    receipt_no = db.Column(db.String(50), unique=True, nullable=False, index=True)
    sale_datetime = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    cashier = db.Column(db.String(100), nullable=False)
    payment_method = db.Column(db.String(20), nullable=False)  # CASH, MOBILE_MONEY, CARD, BANK_TRANSFER
    payment_status = db.Column(db.String(20), default='PAID', nullable=False)  # PAID, PENDING, FAILED, VOIDED
    subtotal_tzs = db.Column(db.Integer, nullable=False)  # TZS in integers
    discount_tzs = db.Column(db.Integer, default=0)  # TZS in integers
    total_tzs = db.Column(db.Integer, nullable=False)  # TZS in integers
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relationships
    items = db.relationship('SaleItem', backref='sale', cascade='all, delete-orphan', lazy='joined')
    
    def to_dict(self):
        """Convert sale to dictionary for API responses"""
        return {
            'sale_id': self.sale_id,
            'receipt_no': self.receipt_no,
            'sale_datetime': self.sale_datetime.isoformat(),
            'cashier': self.cashier,
            'payment_method': self.payment_method,
            'payment_status': self.payment_status,
            'subtotal_tzs': self.subtotal_tzs,
            'discount_tzs': self.discount_tzs,
            'total_tzs': self.total_tzs,
            'items_count': len(self.items),
            'notes': self.notes,
            'items': [item.to_dict() for item in self.items],
            'created_at': self.created_at.isoformat()
        }
    
    def __repr__(self):
        return f'<Sale {self.receipt_no}: {self.total_tzs} TZS>'


class SaleItem(db.Model):
    """
    Sale Item Model - Line items for a sale
    
    Each sale can have multiple items:
    - References SKU but stores snapshot of name and price
    - Snapshot fields ensure historical accuracy
    - Price and name can change, but receipt shows what was actual
    """
    __tablename__ = 'sale_items'
    
    sale_item_id = db.Column(db.Integer, primary_key=True)
    sale_id = db.Column(db.Integer, db.ForeignKey('sales.sale_id'), nullable=False, index=True)
    sku = db.Column(db.String(50), db.ForeignKey('products.sku'), nullable=False)
    product_name_snapshot = db.Column(db.String(255), nullable=False)  # Name at time of sale
    unit_price_tzs_snapshot = db.Column(db.Integer, nullable=False)  # Price at time of sale (TZS integer)
    qty = db.Column(db.Integer, nullable=False)
    line_total_tzs = db.Column(db.Integer, nullable=False)  # qty * unit_price (TZS integer)
    
    # Relationships
    product = db.relationship('Product')
    
    def to_dict(self):
        """Convert sale item to dictionary"""
        return {
            'sale_item_id': self.sale_item_id,
            'sku': self.sku,
            'product_name_snapshot': self.product_name_snapshot,
            'unit_price_tzs_snapshot': self.unit_price_tzs_snapshot,
            'qty': self.qty,
            'line_total_tzs': self.line_total_tzs
        }
    
    def __repr__(self):
        return f'<SaleItem {self.sku} x{self.qty} @ {self.line_total_tzs} TZS>'


# ==================== UTILITY FUNCTIONS ====================

def generate_sku():
    """
    Auto-generate SKU if not provided
    Format: SKU-YYYYMMDDHHMMSS-XX
    Example: SKU-20260304133042-01
    """
    now = datetime.utcnow()
    timestamp = now.strftime('%Y%m%d%H%M%S')
    
    # Get count of products created in current second
    count = Product.query.filter(
        Product.created_at >= now.replace(microsecond=0)
    ).count()
    
    return f'SKU-{timestamp}-{str(count + 1).zfill(2)}'


# ==================== INVENTORY FUNCTIONS ====================

def create_stock_in(sku, quantity, reference=None, notes=None, created_by=None):
    """
    Record goods received from supplier
    
    Args:
        sku: Product SKU
        quantity: Units received
        reference: Supplier invoice number
        notes: Optional notes
        created_by: User recording movement
    
    Returns:
        InventoryMovement object or error
    """
    product = Product.query.filter_by(sku=sku).first()
    if not product:
        return None, f"Product {sku} not found"
    
    movement = InventoryMovement(
        sku=sku,
        movement_type='STOCK_IN',
        quantity=quantity,
        reference=reference,
        notes=notes,
        created_by=created_by
    )
    
    db.session.add(movement)
    db.session.commit()
    return movement, None


def create_adjustment(sku, quantity, reason=None, notes=None, created_by=None):
    """
    Record manual stock adjustment (damage, loss, recount correction)
    
    Args:
        sku: Product SKU
        quantity: Amount to adjust (can be negative)
        reason: DAMAGE, LOSS, RECOUNT, etc
        notes: Explanation
        created_by: User recording movement
    
    Returns:
        InventoryMovement or error
    """
    product = Product.query.filter_by(sku=sku).first()
    if not product:
        return None, f"Product {sku} not found"
    
    current = product.get_current_stock()
    if current + quantity < 0:
        return None, f"Adjustment would result in negative stock (current: {current}, adjustment: {quantity})"
    
    movement = InventoryMovement(
        sku=sku,
        movement_type='ADJUSTMENT',
        quantity=abs(quantity),
        reference=reason or 'MANUAL_ADJUSTMENT',
        notes=notes,
        created_by=created_by
    )
    
    db.session.add(movement)
    db.session.commit()
    return movement, None


def create_sale(sku, quantity, reference=None, created_by=None):
    """
    Record product sale (usually from POS)
    
    Args:
        sku: Product SKU
        quantity: Units sold
        reference: POS receipt or transaction ID
        created_by: Cashier/user name
    
    Returns:
        InventoryMovement or error message
    """
    product = Product.query.filter_by(sku=sku).first()
    if not product:
        return None, f"Product {sku} not found"
    
    current = product.get_current_stock()
    if current < quantity:
        return None, f"Insufficient stock (current: {current}, requested: {quantity})"
    
    movement = InventoryMovement(
        sku=sku,
        movement_type='SALE',
        quantity=quantity,
        reference=reference or 'POS_SALE',
        created_by=created_by
    )
    
    db.session.add(movement)
    db.session.commit()
    return movement, None


def create_return(sku, quantity, reference=None, notes=None, created_by=None):
    """
    Record customer return
    
    Args:
        sku: Product SKU
        quantity: Units returned
        reference: Original receipt/order number
        notes: Return reason
        created_by: User processing return
    
    Returns:
        InventoryMovement or error
    """
    product = Product.query.filter_by(sku=sku).first()
    if not product:
        return None, f"Product {sku} not found"
    
    movement = InventoryMovement(
        sku=sku,
        movement_type='RETURN',
        quantity=quantity,
        reference=reference or 'CUSTOMER_RETURN',
        notes=notes,
        created_by=created_by
    )
    
    db.session.add(movement)
    db.session.commit()
    return movement, None


def get_inventory_summary(sku):
    """
    Get complete inventory picture for a product
    
    Returns dict with:
    - current_stock
    - reorder_level
    - stock_status (Good/Medium/Low)
    - last_movement
    - movement_count
    """
    product = Product.query.filter_by(sku=sku).first()
    if not product:
        return None
    
    current = product.get_current_stock()
    movements = InventoryMovement.query.filter_by(sku=sku).order_by(
        InventoryMovement.created_at.desc()
    ).all()
    
    status = 'Good'
    if current < product.reorder_level:
        status = 'Low Stock'
    elif current < (product.reorder_level * 1.5):
        status = 'Medium Stock'
    
    return {
        'sku': sku,
        'product_name': product.name,
        'current_stock': current,
        'reorder_level': product.reorder_level,
        'status': status,
        'cost_price_tzs': float(product.cost_price_tzs),
        'total_value_tzs': current * float(product.cost_price_tzs),
        'last_movement': movements[0].created_at if movements else None,
        'movement_count': len(movements)
    }


# ==================== POS SALES FUNCTIONS ====================

def generate_receipt_no():
    """
    Generate unique receipt number in format SR-SALE-000001
    SR = Smart Retail, SALE = transaction type, 6-digit counter
    """
    last_sale = Sale.query.order_by(Sale.sale_id.desc()).first()
    next_num = (last_sale.sale_id + 1) if last_sale else 1
    return f'SR-SALE-{str(next_num).zfill(6)}'


def validate_cart_items(items):
    """
    Validate all items in cart before checkout
    
    Args:
        items: List of dicts with 'sku' and 'qty'
    
    Returns:
        (is_valid, error_message, product_data)
    """
    if not items or len(items) == 0:
        return False, "Cart is empty", {}
    
    product_data = {}
    
    for item in items:
        sku = item.get('sku')
        qty = item.get('qty', 0)
        
        if not sku:
            return False, "Missing SKU in cart item", {}
        
        if not isinstance(qty, int) or qty <= 0:
            return False, f"Invalid quantity for {sku}: {qty}", {}
        
        product = Product.query.filter_by(sku=sku).first()
        if not product:
            return False, f"Product {sku} not found", {}
        
        if not product.is_active:
            return False, f"Product {sku} is inactive", {}
        
        current_stock = product.get_current_stock()
        if current_stock < qty:
            return False, f"Insufficient stock for {product.name}: have {current_stock}, need {qty}", {}
        
        product_data[sku] = {
            'product': product,
            'qty': qty,
            'unit_price': int(product.selling_price_tzs),
            'product_name': product.name,
            'line_total': int(product.selling_price_tzs) * qty
        }
    
    return True, None, product_data


def checkout_sale(cashier, payment_method, items, discount_tzs=0, notes=None):
    """
    Complete a POS sale - ATOMIC TRANSACTION
    
    Args:
        cashier: Cashier name/ID
        payment_method: CASH, MOBILE_MONEY, CARD, BANK_TRANSFER
        items: List of {'sku': str, 'qty': int}
        discount_tzs: Discount amount (TZS integer)
        notes: Optional notes
    
    Returns:
        (sale, error_message) or (None, error_message)
    """
    
    # Validate items
    is_valid, error, product_data = validate_cart_items(items)
    if not is_valid:
        return None, error
    
    try:
        # Calculate totals
        subtotal_tzs = sum(item['line_total'] for item in product_data.values())
        discount_tzs = max(0, min(discount_tzs, subtotal_tzs))  # Clamp discount to subtotal
        total_tzs = subtotal_tzs - discount_tzs
        
        # Generate receipt number
        receipt_no = generate_receipt_no()
        
        # Create sale header (within transaction)
        sale = Sale(
            receipt_no=receipt_no,
            cashier=cashier,
            payment_method=payment_method,
            payment_status='PAID',
            subtotal_tzs=subtotal_tzs,
            discount_tzs=discount_tzs,
            total_tzs=total_tzs,
            notes=notes
        )
        db.session.add(sale)
        db.session.flush()  # Get sale_id without committing
        
        # Create sale items
        for sku, data in product_data.items():
            sale_item = SaleItem(
                sale_id=sale.sale_id,
                sku=sku,
                product_name_snapshot=data['product_name'],
                unit_price_tzs_snapshot=data['unit_price'],
                qty=data['qty'],
                line_total_tzs=data['line_total']
            )
            db.session.add(sale_item)
        
        # Create inventory movements (SALE type with negative qty)
        for sku, data in product_data.items():
            movement = InventoryMovement(
                sku=sku,
                movement_type='SALE',
                quantity=data['qty'],
                reference=receipt_no,
                created_by=cashier
            )
            db.session.add(movement)
        
        # Commit all or nothing
        db.session.commit()
        
        return sale, None
        
    except Exception as e:
        db.session.rollback()
        return None, f"Checkout failed: {str(e)}"


def void_sale(sale_id, reason=None, user=None):
    """
    Void a sale (reverse stock, mark as VOIDED)
    
    Audit-safe: doesn't delete, creates RETURN movements to reverse
    
    Args:
        sale_id: Sale ID to void
        reason: Why sale was voided
        user: User performing void
    
    Returns:
        (sale, error_message) or (None, error_message)
    """
    sale = Sale.query.get(sale_id)
    if not sale:
        return None, f"Sale {sale_id} not found"
    
    if sale.payment_status == 'VOIDED':
        return None, f"Sale {sale.receipt_no} is already voided"
    
    try:
        # Create RETURN movements to reverse each sale item
        for item in sale.items:
            movement = InventoryMovement(
                sku=item.sku,
                movement_type='RETURN',
                quantity=item.qty,
                reference=f"{sale.receipt_no}-VOID",
                notes=f"Void reason: {reason}" if reason else "Sale voided",
                created_by=user
            )
            db.session.add(movement)
        
        # Mark sale as voided
        sale.payment_status = 'VOIDED'
        sale.updated_at = datetime.utcnow()
        
        db.session.commit()
        return sale, None
        
    except Exception as e:
        db.session.rollback()
        return None, f"Void failed: {str(e)}"


def init_db():
    """Initialize database with schema"""
    with app.app_context():
        db.create_all()



def seed_sample_data():
    """Seed database with Tanzanian retail example products"""
    sample_products = [
        {
            'sku': 'SOAP001',
            'name': 'Lux White Soap 100g',
            'category': 'Personal Care',
            'cost_price_tzs': Decimal('1200'),
            'selling_price_tzs': Decimal('1500'),
            'quantity_in_stock': 45,
            'reorder_level': 50,
            'description': 'Premium white soap for daily use',
            'image_url': 'https://via.placeholder.com/200?text=Lux+Soap'
        },
        {
            'sku': 'RICE001',
            'name': 'Pishori Rice 2kg',
            'category': 'Grains',
            'cost_price_tzs': Decimal('3500'),
            'selling_price_tzs': Decimal('4200'),
            'quantity_in_stock': 85,
            'reorder_level': 30,
            'description': 'Premium basmati-style rice',
            'image_url': 'https://via.placeholder.com/200?text=Rice'
        },
        {
            'sku': 'OIL001',
            'name': 'Sunflower Oil 1L',
            'category': 'Cooking',
            'cost_price_tzs': Decimal('5500'),
            'selling_price_tzs': Decimal('6800'),
            'quantity_in_stock': 12,
            'reorder_level': 20,
            'description': 'Pure refined sunflower oil',
            'image_url': 'https://via.placeholder.com/200?text=Oil'
        },
        {
            'sku': 'FLOUR001',
            'name': 'PRO Wheat Flour 2kg',
            'category': 'Grains',
            'cost_price_tzs': Decimal('2800'),
            'selling_price_tzs': Decimal('3500'),
            'quantity_in_stock': 75,
            'reorder_level': 40,
            'description': 'PRO flour for baking and cooking',
            'image_url': 'https://via.placeholder.com/200?text=Flour'
        },
        {
            'sku': 'SALT001',
            'name': 'Everta Salt 500g',
            'category': 'Spices',
            'cost_price_tzs': Decimal('500'),
            'selling_price_tzs': Decimal('800'),
            'quantity_in_stock': 120,
            'reorder_level': 100,
            'description': 'Iodized salt with anti-caking agent',
            'image_url': 'https://via.placeholder.com/200?text=Salt'
        },
        {
            'sku': 'SUGAR001',
            'name': 'Sukari Refined Sugar 1kg',
            'category': 'Sweeteners',
            'cost_price_tzs': Decimal('2200'),
            'selling_price_tzs': Decimal('2800'),
            'quantity_in_stock': 22,
            'reorder_level': 30,
            'description': 'Pure white refined sugar',
            'image_url': 'https://via.placeholder.com/200?text=Sugar'
        },
        {
            'sku': 'MILK001',
            'name': 'Ushindi Milk Powder 400g',
            'category': 'Dairy',
            'cost_price_tzs': Decimal('7500'),
            'selling_price_tzs': Decimal('9500'),
            'quantity_in_stock': 60,
            'reorder_level': 25,
            'description': 'Fortified full cream milk powder',
            'image_url': 'https://via.placeholder.com/200?text=Milk'
        },
        {
            'sku': 'TEA001',
            'name': 'Brooke Bond Tanganyika Tea 50g',
            'category': 'Beverages',
            'cost_price_tzs': Decimal('6500'),
            'selling_price_tzs': Decimal('8000'),
            'quantity_in_stock': 55,
            'reorder_level': 35,
            'description': 'Premium loose leaf black tea',
            'image_url': 'https://via.placeholder.com/200?text=Tea'
        },
        {
            'sku': 'COKE001',
            'name': 'Coca Cola 400ml Bottle',
            'category': 'Soft Drinks',
            'cost_price_tzs': Decimal('1400'),
            'selling_price_tzs': Decimal('2000'),
            'quantity_in_stock': 150,
            'reorder_level': 60,
            'description': 'Carbonated soft drink',
            'image_url': 'https://via.placeholder.com/200?text=CocaCola'
        },
        {
            'sku': 'JUICE001',
            'name': 'Minute Maid Orange Juice 200ml',
            'category': 'Beverages',
            'cost_price_tzs': Decimal('1200'),
            'selling_price_tzs': Decimal('1800'),
            'quantity_in_stock': 35,
            'reorder_level': 50,
            'description': 'Orange juice concentrate',
            'image_url': 'https://via.placeholder.com/200?text=Juice'
        }
    ]
    
    with app.app_context():
        for prod_data in sample_products:
            if not Product.query.filter_by(sku=prod_data['sku']).first():
                product = Product(**prod_data)
                db.session.add(product)
        db.session.commit()
        
        # Create initial inventory movements for sample products
        initial_movements = [
            # SOAP001: Received 50, sold 3 twice, damaged 1 = current 45
            ('SOAP001', 'STOCK_IN', 50, 'SUP-001'),
            ('SOAP001', 'SALE', 3, 'POS-001'),
            ('SOAP001', 'SALE', 2, 'POS-002'),
            
            # RICE001: Received 100, sold various = current 85
            ('RICE001', 'STOCK_IN', 100, 'SUP-002'),
            ('RICE001', 'SALE', 8, 'POS-003'),
            ('RICE001', 'SALE', 7, 'POS-004'),
            
            # OIL001: Low stock - received 20, sold heavily = current 12
            ('OIL001', 'STOCK_IN', 20, 'SUP-003'),
            ('OIL001', 'SALE', 5, 'POS-005'),
            ('OIL001', 'SALE', 3, 'POS-006'),
            
            # FLOUR001: Received 100, sold = current 75
            ('FLOUR001', 'STOCK_IN', 100, 'SUP-004'),
            ('FLOUR001', 'SALE', 25, 'POS-007'),
            
            # SALT001: High stock - received 150, low sales
            ('SALT001', 'STOCK_IN', 150, 'SUP-005'),
            ('SALT001', 'SALE', 30, 'POS-008'),
            
            # SUGAR001: Low stock - received 50, sold heavily
            ('SUGAR001', 'STOCK_IN', 50, 'SUP-006'),
            ('SUGAR001', 'SALE', 20, 'POS-009'),
            ('SUGAR001', 'ADJUSTMENT', 8, 'RECOUNT-01'),
            
            # MILK001: Received 80, sold = current 60
            ('MILK001', 'STOCK_IN', 80, 'SUP-007'),
            ('MILK001', 'SALE', 20, 'POS-010'),
            
            # TEA001: Received 70, sold = current 55
            ('TEA001', 'STOCK_IN', 70, 'SUP-008'),
            ('TEA001', 'SALE', 15, 'POS-011'),
            
            # COKE001: High demand - received 200, heavy sales
            ('COKE001', 'STOCK_IN', 200, 'SUP-009'),
            ('COKE001', 'SALE', 30, 'POS-012'),
            ('COKE001', 'SALE', 20, 'POS-013'),
            
            # JUICE001: Received 50, sold = current 35
            ('JUICE001', 'STOCK_IN', 50, 'SUP-010'),
            ('JUICE001', 'SALE', 15, 'POS-014'),
        ]
        
        for sku, move_type, qty, ref in initial_movements:
            product = Product.query.filter_by(sku=sku).first()
            if product and InventoryMovement.query.filter_by(sku=sku).first() is None:
                movement = InventoryMovement(
                    sku=sku,
                    movement_type=move_type,
                    quantity=qty,
                    reference=ref,
                    created_by='SYSTEM'
                )
                db.session.add(movement)
        
        db.session.commit()


def allowed_file(filename):
    """Check if file is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


def get_sample_images():
    """Sample image URLs for products"""
    images = {
        'personal care': [
            'https://via.placeholder.com/200?text=Soap',
            'https://via.placeholder.com/200?text=Shampoo',
            'https://via.placeholder.com/200?text=Toothpaste'
        ],
        'grains': [
            'https://via.placeholder.com/200?text=Rice',
            'https://via.placeholder.com/200?text=Flour',
            'https://via.placeholder.com/200?text=Maize'
        ],
        'cooking': [
            'https://via.placeholder.com/200?text=Cooking+Oil',
            'https://via.placeholder.com/200?text=Salt',
            'https://via.placeholder.com/200?text=Spices'
        ],
        'spices': [
            'https://via.placeholder.com/200?text=Spices',
            'https://via.placeholder.com/200?text=Pepper',
            'https://via.placeholder.com/200?text=Garlic'
        ],
        'beverages': [
            'https://via.placeholder.com/200?text=Soda',
            'https://via.placeholder.com/200?text=Juice',
            'https://via.placeholder.com/200?text=Water'
        ],
        'dairy': [
            'https://via.placeholder.com/200?text=Milk',
            'https://via.placeholder.com/200?text=Cheese',
            'https://via.placeholder.com/200?text=Yogurt'
        ]
    }
    return images


def parse_csv_file(file_path):
    """Parse CSV file and return products list"""
    products = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if not row or not row.get('name'):
                    continue
                products.append(row)
    except Exception as e:
        return None, str(e)
    return products, None


def parse_excel_file(file_path):
    """Parse Excel file and return products list"""
    products = []
    try:
        if not OPENPYXL_AVAILABLE:
            return None, "openpyxl not installed"
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Parse data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Skip empty rows
                continue
            
            product_data = {}
            for col_idx, header in enumerate(headers):
                if header and col_idx < len(row):
                    product_data[header] = row[col_idx]
            
            if product_data.get('name'):
                products.append(product_data)
    
    except Exception as e:
        return None, str(e)
    return products, None


def validate_product_data(product_data, image_map):
    """Validate and prepare product data for database"""
    errors = []
    
    # Check required fields
    if not product_data.get('name'):
        errors.append('name')
    if not product_data.get('category'):
        errors.append('category')
    if not product_data.get('cost_price_tzs'):
        errors.append('cost_price_tzs')
    if not product_data.get('selling_price_tzs'):
        errors.append('selling_price_tzs')
    
    if errors:
        return None, f"Missing fields: {', '.join(errors)}"
    
    try:
        # Assign sample image if not provided
        image_url = product_data.get('image_url')
        if not image_url:
            category_lower = str(product_data.get('category', '')).lower()
            sample_images = get_sample_images()
            images_for_category = sample_images.get(category_lower, sample_images['personal care'])
            image_idx = len(image_map) % len(images_for_category)
            image_url = images_for_category[image_idx]
        
        prepared = {
            'sku': product_data.get('sku') or generate_sku(),
            'name': str(product_data['name']).strip(),
            'category': str(product_data['category']).strip(),
            'cost_price_tzs': Decimal(str(product_data['cost_price_tzs'])),
            'selling_price_tzs': Decimal(str(product_data['selling_price_tzs'])),
            'reorder_level': int(product_data.get('reorder_level', 10)),
            'description': product_data.get('description', ''),
            'image_url': image_url,
            'is_active': True
        }
        return prepared, None
    except Exception as e:
        return None, f"Invalid data: {str(e)}"


# ==================== REST API ENDPOINTS ====================

@app.route('/api/v1/products', methods=['GET'])
def api_get_products():
    """Get all products with pagination and filtering"""
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 20, type=int)
    category = request.args.get('category', None)
    is_active = request.args.get('is_active', 'true').lower() == 'true'
    
    query = Product.query.filter_by(is_active=is_active) if is_active else Product.query
    
    if category:
        query = query.filter_by(category=category)
    
    pagination = query.paginate(page=page, per_page=limit, error_out=False)
    
    return jsonify({
        'status': 'success',
        'data': [p.to_dict() for p in pagination.items],
        'pagination': {
            'total': pagination.total,
            'page': page,
            'limit': limit,
            'total_pages': pagination.pages
        }
    }), 200


@app.route('/api/v1/products/<int:product_id>', methods=['GET'])
def api_get_product(product_id):
    """Get single product by ID"""
    product = Product.query.get_or_404(product_id)
    return jsonify({
        'status': 'success',
        'data': product.to_dict()
    }), 200


@app.route('/api/v1/products', methods=['POST'])
def api_create_product():
    """Create new product"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['name', 'category', 'cost_price_tzs', 'selling_price_tzs']
        for field in required_fields:
            if field not in data:
                return jsonify({'status': 'error', 'message': f'Missing field: {field}'}), 400
        
        # Generate SKU if not provided
        sku = data.get('sku') or generate_sku()
        
        # Check for duplicate SKU
        if Product.query.filter_by(sku=sku).first():
            return jsonify({'status': 'error', 'message': 'SKU already exists'}), 409
        
        product = Product(
            sku=sku,
            name=data['name'],
            category=data['category'],
            cost_price_tzs=Decimal(str(data['cost_price_tzs'])),
            selling_price_tzs=Decimal(str(data['selling_price_tzs'])),
            reorder_level=data.get('reorder_level', 10),
            description=data.get('description'),
            image_url=data.get('image_url'),
            is_active=data.get('is_active', True)
        )
        
        db.session.add(product)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Product created',
            'data': product.to_dict()
        }), 201
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400


@app.route('/api/v1/products/<int:product_id>', methods=['PUT'])
def api_update_product(product_id):
    """Update existing product"""
    try:
        product = Product.query.get_or_404(product_id)
        data = request.get_json()
        
        # Update fields
        if 'name' in data:
            product.name = data['name']
        if 'category' in data:
            product.category = data['category']
        if 'cost_price_tzs' in data:
            product.cost_price_tzs = Decimal(str(data['cost_price_tzs']))
        if 'selling_price_tzs' in data:
            product.selling_price_tzs = Decimal(str(data['selling_price_tzs']))
        if 'reorder_level' in data:
            product.reorder_level = data['reorder_level']
        if 'description' in data:
            product.description = data['description']
        if 'image_url' in data:
            product.image_url = data['image_url']
        if 'is_active' in data:
            product.is_active = data['is_active']
        
        product.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Product updated',
            'data': product.to_dict()
        }), 200
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400


@app.route('/api/v1/products/<int:product_id>', methods=['DELETE'])
def api_delete_product(product_id):
    """Soft-delete product (archive)"""
    try:
        product = Product.query.get_or_404(product_id)
        product.is_active = False
        product.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Product archived'
        }), 200
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400


@app.route('/api/v1/products/search', methods=['GET'])
def api_search_products():
    """Search products by name or SKU"""
    query = request.args.get('q', '').strip()
    
    if not query or len(query) < 2:
        return jsonify({'status': 'error', 'message': 'Search query too short'}), 400
    
    products = Product.query.filter(
        (Product.name.ilike(f'%{query}%')) | 
        (Product.sku.ilike(f'%{query}%'))
    ).filter_by(is_active=True).limit(50).all()
    
    return jsonify({
        'status': 'success',
        'data': [p.to_dict() for p in products]
    }), 200


@app.route('/api/v1/products/category/<category>', methods=['GET'])
def api_get_by_category(category):
    """Get products by category"""
    products = Product.query.filter_by(
        category=category,
        is_active=True
    ).all()
    
    return jsonify({
        'status': 'success',
        'data': [p.to_dict() for p in products]
    }), 200


@app.route('/api/v1/products/low-stock', methods=['GET'])
def api_low_stock_products():
    """Get products below reorder level"""
    products = Product.query.filter(
        Product.reorder_level > 0
    ).filter_by(is_active=True).all()
    
    # In Phase 2, integrate with inventory table to check actual stock
    # For now, return products with low reorder thresholds
    
    return jsonify({
        'status': 'success',
        'data': [p.to_dict() for p in products[:10]]
    }), 200


@app.route('/api/v1/products/stats', methods=['GET'])
def api_product_stats():
    """Get product statistics"""
    total = Product.query.count()
    active = Product.query.filter_by(is_active=True).count()
    
    products = Product.query.all()
    total_cost = sum(float(p.cost_price_tzs) for p in products)
    total_selling = sum(float(p.selling_price_tzs) for p in products)
    
    # Calculate average margin
    margins = []
    for p in products:
        cost = float(p.cost_price_tzs)
        selling = float(p.selling_price_tzs)
        if cost > 0:
            margins.append((selling - cost) / cost * 100)
    avg_margin = sum(margins) / len(margins) if margins else 0
    
    return jsonify({
        'status': 'success',
        'data': {
            'total_products': total,
            'active_products': active,
            'archived_products': total - active,
            'total_cost_value_tzs': int(total_cost),
            'total_selling_value_tzs': int(total_selling),
            'average_margin_percent': round(avg_margin, 2)
        }
    }), 200


@app.route('/api/v1/categories', methods=['GET'])
def api_get_categories():
    """Get list of all unique categories"""
    categories = db.session.query(Product.category).distinct().filter_by(
        is_active=True
    ).all()
    
    return jsonify({
        'status': 'success',
        'data': [cat[0] for cat in categories]
    }), 200


# ==================== INVENTORY API ENDPOINTS ====================

@app.route('/api/v1/inventory/stock/<sku>', methods=['GET'])
def api_get_inventory(sku):
    """Get inventory summary for a product"""
    summary = get_inventory_summary(sku)
    
    if not summary:
        return jsonify({'status': 'error', 'message': 'Product not found'}), 404
    
    return jsonify({
        'status': 'success',
        'data': summary
    }), 200


@app.route('/api/v1/inventory/stock-in', methods=['POST'])
def api_stock_in():
    """Record goods received from supplier"""
    data = request.get_json()
    
    sku = data.get('sku')
    quantity = data.get('quantity')
    reference = data.get('reference')  # Supplier invoice
    notes = data.get('notes')
    created_by = data.get('created_by')
    
    if not sku or not quantity:
        return jsonify({'status': 'error', 'message': 'SKU and quantity required'}), 400
    
    movement, error = create_stock_in(sku, quantity, reference, notes, created_by)
    
    if error:
        return jsonify({'status': 'error', 'message': error}), 400
    
    return jsonify({
        'status': 'success',
        'message': f'Recorded {quantity} units for {sku}',
        'data': movement.to_dict()
    }), 201


@app.route('/api/v1/inventory/adjust', methods=['POST'])
def api_adjust_stock():
    """Record manual stock adjustment"""
    data = request.get_json()
    
    sku = data.get('sku')
    quantity = data.get('quantity')  # Can be negative
    reason = data.get('reason')  # DAMAGE, LOSS, RECOUNT
    notes = data.get('notes')
    created_by = data.get('created_by')
    
    if not sku or quantity is None:
        return jsonify({'status': 'error', 'message': 'SKU and quantity required'}), 400
    
    movement, error = create_adjustment(sku, quantity, reason, notes, created_by)
    
    if error:
        return jsonify({'status': 'error', 'message': error}), 400
    
    return jsonify({
        'status': 'success',
        'message': f'Adjustment recorded for {sku}',
        'data': movement.to_dict()
    }), 201


@app.route('/api/v1/inventory/sale', methods=['POST'])
def api_sale():
    """Record product sale (from POS)"""
    data = request.get_json()
    
    sku = data.get('sku')
    quantity = data.get('quantity')
    reference = data.get('reference')  # POS receipt
    created_by = data.get('created_by')  # Cashier
    
    if not sku or not quantity:
        return jsonify({'status': 'error', 'message': 'SKU and quantity required'}), 400
    
    movement, error = create_sale(sku, quantity, reference, created_by)
    
    if error:
        return jsonify({'status': 'error', 'message': error}), 400
    
    return jsonify({
        'status': 'success',
        'message': f'Sale recorded: {quantity} units of {sku}',
        'data': movement.to_dict()
    }), 201


@app.route('/api/v1/inventory/return', methods=['POST'])
def api_return_product():
    """Record customer return"""
    data = request.get_json()
    
    sku = data.get('sku')
    quantity = data.get('quantity')
    reference = data.get('reference')  # Original receipt
    notes = data.get('notes')  # Return reason
    created_by = data.get('created_by')
    
    if not sku or not quantity:
        return jsonify({'status': 'error', 'message': 'SKU and quantity required'}), 400
    
    movement, error = create_return(sku, quantity, reference, notes, created_by)
    
    if error:
        return jsonify({'status': 'error', 'message': error}), 400
    
    return jsonify({
        'status': 'success',
        'message': f'Return recorded: {quantity} units of {sku}',
        'data': movement.to_dict()
    }), 201


@app.route('/api/v1/inventory/movements/<sku>', methods=['GET'])
def api_get_movements(sku):
    """Get movement history for a product"""
    movements = InventoryMovement.query.filter_by(sku=sku).order_by(
        InventoryMovement.created_at.desc()
    ).all()
    
    return jsonify({
        'status': 'success',
        'sku': sku,
        'total_movements': len(movements),
        'data': [m.to_dict() for m in movements]
    }), 200


@app.route('/api/v1/inventory/all', methods=['GET'])
def api_get_all_inventory():
    """Get inventory summary for all products"""
    products = Product.query.filter_by(is_active=True).all()
    
    summaries = []
    for product in products:
        summary = get_inventory_summary(product.sku)
        if summary:
            summaries.append(summary)
    
    return jsonify({
        'status': 'success',
        'total_products': len(summaries),
        'data': summaries
    }), 200


@app.route('/api/v1/inventory/movements/all', methods=['GET'])
def api_get_all_movements():
    """Get all inventory movements across all products"""
    movements = InventoryMovement.query.order_by(
        InventoryMovement.created_at.desc()
    ).all()
    
    movement_data = []
    for movement in movements:
        product = Product.query.filter_by(sku=movement.sku).first()
        movement_data.append({
            'movement_id': movement.movement_id,
            'sku': movement.sku,
            'product_name': product.name if product else 'Unknown',
            'movement_type': movement.movement_type,
            'quantity': movement.quantity,
            'reference': movement.reference,
            'notes': movement.notes,
            'created_by': movement.created_by,
            'created_at': movement.created_at.isoformat()
        })
    
    return jsonify({
        'status': 'success',
        'total_movements': len(movement_data),
        'data': movement_data
    }), 200


# ==================== POS SALES API ENDPOINTS ====================

@app.route('/api/pos/products', methods=['GET'])
def api_pos_products():
    """
    POS product search endpoint
    
    Query params:
    - q: search by name/SKU
    - category: filter by category
    - limit: max results (default 50)
    """
    q = request.args.get('q', '').strip()
    category = request.args.get('category', '').strip()
    limit = min(int(request.args.get('limit', 50)), 500)
    
    query = Product.query.filter_by(is_active=True)
    
    if q:
        query = query.filter(
            db.or_(
                Product.sku.ilike(f'%{q}%'),
                Product.name.ilike(f'%{q}%')
            )
        )
    
    if category:
        query = query.filter_by(category=category)
    
    products = query.limit(limit).all()
    
    data = []
    for product in products:
        current_stock = product.get_current_stock()
        data.append({
            'sku': product.sku,
            'name': product.name,
            'category': product.category,
            'selling_price_tzs': int(product.selling_price_tzs),
            'current_stock': current_stock,
            'is_in_stock': current_stock > 0,
            'image_url': product.image_url
        })
    
    return jsonify({
        'status': 'success',
        'count': len(data),
        'data': data
    }), 200


@app.route('/api/pos/sales', methods=['POST'])
def api_pos_checkout():
    """
    POS checkout endpoint - creates sale with all items atomically
    
    Request body:
    {
        "cashier": "cashier1",
        "payment_method": "CASH",
        "discount_tzs": 0,
        "items": [
            {"sku": "SOAP001", "qty": 2},
            {"sku": "RICE001", "qty": 1}
        ],
        "notes": "optional notes"
    }
    
    Returns receipt_no and sale_id
    """
    data = request.get_json() or {}
    
    cashier = data.get('cashier', 'Unknown')
    payment_method = data.get('payment_method', 'CASH')
    discount_tzs = int(data.get('discount_tzs', 0))
    items = data.get('items', [])
    notes = data.get('notes')
    
    # Validate input
    if not payment_method in ['CASH', 'MOBILE_MONEY', 'CARD', 'BANK_TRANSFER']:
        return jsonify({
            'status': 'error',
            'message': 'Invalid payment method'
        }), 400
    
    # Checkout
    sale, error = checkout_sale(cashier, payment_method, items, discount_tzs, notes)
    
    if error:
        return jsonify({
            'status': 'error',
            'message': error
        }), 400
    
    return jsonify({
        'status': 'success',
        'sale_id': sale.sale_id,
        'receipt_no': sale.receipt_no,
        'total_tzs': sale.total_tzs,
        'payment_status': sale.payment_status
    }), 201


@app.route('/api/pos/sales', methods=['GET'])
def api_pos_sales_list():
    """
    List sales with optional filtering
    
    Query params:
    - from_date: YYYY-MM-DD
    - to_date: YYYY-MM-DD
    - payment_method: filter by method
    - limit: max results (default 50)
    """
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    payment_method = request.args.get('payment_method')
    limit = min(int(request.args.get('limit', 50)), 500)
    
    query = Sale.query
    
    if from_date:
        try:
            from_dt = datetime.strptime(from_date, '%Y-%m-%d')
            query = query.filter(Sale.sale_datetime >= from_dt)
        except ValueError:
            pass
    
    if to_date:
        try:
            to_dt = datetime.strptime(to_date, '%Y-%m-%d')
            to_dt = to_dt.replace(hour=23, minute=59, second=59)
            query = query.filter(Sale.sale_datetime <= to_dt)
        except ValueError:
            pass
    
    if payment_method:
        query = query.filter_by(payment_method=payment_method)
    
    sales = query.order_by(Sale.sale_datetime.desc()).limit(limit).all()
    
    data = [{
        'sale_id': s.sale_id,
        'receipt_no': s.receipt_no,
        'sale_datetime': s.sale_datetime.isoformat(),
        'cashier': s.cashier,
        'payment_method': s.payment_method,
        'payment_status': s.payment_status,
        'total_tzs': s.total_tzs,
        'items_count': len(s.items)
    } for s in sales]
    
    return jsonify({
        'status': 'success',
        'count': len(data),
        'data': data
    }), 200


@app.route('/api/pos/sales/<int:sale_id>', methods=['GET'])
def api_pos_sale_detail(sale_id):
    """Get full sale details with all items"""
    sale = Sale.query.get(sale_id)
    if not sale:
        return jsonify({
            'status': 'error',
            'message': f'Sale {sale_id} not found'
        }), 404
    
    return jsonify({
        'status': 'success',
        'data': sale.to_dict()
    }), 200


@app.route('/api/pos/sales/<int:sale_id>/void', methods=['POST'])
def api_pos_void_sale(sale_id):
    """Void a sale - creates RETURN movements to reverse stock"""
    data = request.get_json() or {}
    reason = data.get('reason')
    user = data.get('user', 'Unknown')
    
    sale, error = void_sale(sale_id, reason, user)
    
    if error:
        return jsonify({
            'status': 'error',
            'message': error
        }), 400
    
    return jsonify({
        'status': 'success',
        'message': f'Sale {sale.receipt_no} voided successfully',
        'sale_id': sale.sale_id,
        'receipt_no': sale.receipt_no
    }), 200


# ==================== WEB UI ROUTES ====================

@app.route('/')
def index():
    """Dashboard home page"""
    total_products = Product.query.filter_by(is_active=True).count()
    categories = db.session.query(Product.category).distinct().filter_by(
        is_active=True
    ).count()
    
    return render_template('index.html', 
                         total_products=total_products,
                         total_categories=categories)


@app.route('/products')
def products_list():
    """Products management page"""
    products = Product.query.filter_by(is_active=True).all()
    return render_template('products/list.html', products=products)


@app.route('/products/new', methods=['GET', 'POST'])
def create_product():
    """Create new product"""
    if request.method == 'POST':
        try:
            product = Product(
                sku=request.form.get('sku') or generate_sku(),
                name=request.form.get('name'),
                category=request.form.get('category'),
                cost_price_tzs=Decimal(request.form.get('cost_price_tzs')),
                selling_price_tzs=Decimal(request.form.get('selling_price_tzs')),
                reorder_level=int(request.form.get('reorder_level', 10)),
                description=request.form.get('description'),
                image_url=request.form.get('image_url'),
                is_active=True
            )
            
            db.session.add(product)
            db.session.commit()
            
            flash(f'Product "{product.name}" created successfully', 'success')
            return redirect(url_for('products_list'))
            
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
    
    return render_template('products/form.html', product=None)


@app.route('/products/<int:product_id>', methods=['GET'])
def view_product(product_id):
    """View product details"""
    product = Product.query.get_or_404(product_id)
    return render_template('products/view.html', product=product)


@app.route('/products/<int:product_id>/edit', methods=['GET', 'POST'])
def edit_product(product_id):
    """Edit product"""
    product = Product.query.get_or_404(product_id)
    
    if request.method == 'POST':
        try:
            product.name = request.form.get('name')
            product.category = request.form.get('category')
            product.cost_price_tzs = Decimal(request.form.get('cost_price_tzs'))
            product.selling_price_tzs = Decimal(request.form.get('selling_price_tzs'))
            product.reorder_level = int(request.form.get('reorder_level'))
            product.description = request.form.get('description')
            product.image_url = request.form.get('image_url')
            product.updated_at = datetime.utcnow()
            
            db.session.commit()
            
            flash(f'Product "{product.name}" updated successfully', 'success')
            return redirect(url_for('products_list'))
            
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
    
    return render_template('products/form.html', product=product)


@app.route('/products/<int:product_id>/delete', methods=['POST'])
def delete_product(product_id):
    """Delete (archive) product"""
    try:
        product = Product.query.get_or_404(product_id)
        product.is_active = False
        db.session.commit()
        
        flash(f'Product "{product.name}" archived', 'success')
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('products_list'))


@app.route('/products/bulk/import', methods=['GET', 'POST'])
def bulk_import():
    """Bulk import products from CSV/Excel file"""
    if request.method == 'POST':
        # Check if file is provided
        if 'file' not in request.files:
            flash('No file provided', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(request.url)
        
        if not allowed_file(file.filename):
            flash('Invalid file type. Allowed: CSV, XLSX, XLS', 'danger')
            return redirect(request.url)
        
        try:
            # Create uploads folder if not exists
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # Parse file
            products = None
            error = None
            
            if filename.endswith('.csv'):
                products, error = parse_csv_file(filepath)
            else:  # Excel file
                products, error = parse_excel_file(filepath)
            
            if error:
                flash(f'Error parsing file: {error}', 'danger')
                return redirect(request.url)
            
            if not products:
                flash('No valid products found in file', 'warning')
                return redirect(request.url)
            
            # Validate and import products
            imported = 0
            skipped = 0
            errors_list = []
            image_map = []
            
            for idx, prod_data in enumerate(products, 1):
                prepared, error = validate_product_data(prod_data, image_map)
                
                if error:
                    skipped += 1
                    errors_list.append(f"Row {idx}: {error}")
                    continue
                
                # Check if SKU already exists
                if Product.query.filter_by(sku=prepared['sku']).first():
                    skipped += 1
                    errors_list.append(f"Row {idx}: SKU {prepared['sku']} already exists")
                    continue
                
                try:
                    product = Product(**prepared)
                    db.session.add(product)
                    image_map.append(prepared['sku'])
                    imported += 1
                except Exception as e:
                    skipped += 1
                    errors_list.append(f"Row {idx}: {str(e)}")
            
            # Commit all products
            if imported > 0:
                db.session.commit()
                flash(f'Successfully imported {imported} products!', 'success')
            
            if errors_list and len(errors_list) <= 10:
                for error_msg in errors_list:
                    flash(error_msg, 'warning')
            elif errors_list:
                flash(f'Skipped {skipped} products due to errors', 'warning')
            
            # Clean up uploaded file
            os.remove(filepath)
            
            return redirect(url_for('products_list'))
            
        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'danger')
            return redirect(request.url)
    
    return render_template('products/bulk_import.html')


@app.route('/products/bulk/template')
def download_template():
    """Download CSV template for bulk import with 30 sample products"""
    import io
    import os
    from flask import send_file
    
    # Try to read the COMPLETE_PRODUCT_TEMPLATE.csv file
    template_path = os.path.join('sample_imports', 'COMPLETE_PRODUCT_TEMPLATE.csv')
    
    if os.path.exists(template_path):
        # Send the complete template with 30 products
        return send_file(
            template_path,
            mimetype='text/csv',
            as_attachment=True,
            download_name='product_import_template.csv'
        )
    else:
        # Fallback: Create minimal template if file doesn't exist
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow([
            'sku',
            'name',
            'category',
            'cost_price_tzs',
            'selling_price_tzs',
            'reorder_level',
            'description',
            'image_url'
        ])
        
        # Write sample row
        writer.writerow([
            'PROD001',
            'Sample Product',
            'Personal Care',
            '1000',
            '1500',
            '10',
            'Sample product description',
            'https://via.placeholder.com/200?text=Product'
        ])
        
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode()),
            mimetype='text/csv',
            as_attachment=True,
            download_name='product_import_template.csv'
        )


# ==================== INVENTORY WEB ROUTES ====================

@app.route('/inventory/overview')
def inventory_overview():
    """Inventory management dashboard"""
    return render_template('inventory/overview.html')


@app.route('/inventory/stock-in', methods=['GET', 'POST'])
def inventory_stock_in():
    """Record stock in (supplier goods)"""
    return render_template('inventory/stock_in.html')


@app.route('/inventory/adjust', methods=['GET', 'POST'])
def inventory_adjust():
    """Adjust stock manually"""
    return render_template('inventory/adjust.html')


@app.route('/inventory/history')
@app.route('/inventory/history/<sku>')
def inventory_history(sku=None):
    """View movement history for products"""
    return render_template('inventory/history.html', sku=sku)


# ==================== POS SALES WEB ROUTES ====================

@app.route('/pos/terminal')
def pos_terminal():
    """POS cashier terminal - main sales interface"""
    return render_template('pos/terminal.html')


@app.route('/pos/receipt/<int:sale_id>')
def pos_receipt(sale_id):
    """Receipt view page for printing"""
    sale = Sale.query.get_or_404(sale_id)
    return render_template('pos/receipt.html', sale=sale)


@app.route('/pos/sales')
def pos_sales_history():
    """Sales history for management review"""
    return render_template('pos/sales_history.html')


# ==================== ERROR HANDLERS ====================

# ==================== REPORTS & ANALYTICS ENDPOINTS ====================

@app.route('/api/reports/dashboard', methods=['GET'])
def api_reports_dashboard():
    """
    Dashboard KPI endpoint
    
    Query params:
    - date: YYYY-MM-DD (defaults to today)
    
    Returns KPIs for today, this week, this month
    """
    date_str = request.args.get('date', datetime.utcnow().strftime('%Y-%m-%d'))
    
    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        target_date = datetime.utcnow()
    
    # Today's KPIs
    today_start = target_date.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start.replace(hour=23, minute=59, second=59)
    
    today_sales = Sale.query.filter(
        Sale.payment_status == 'PAID',
        Sale.sale_datetime >= today_start,
        Sale.sale_datetime <= today_end
    ).all()
    
    today_total = sum(s.total_tzs for s in today_sales)
    today_count = len(today_sales)
    today_avg = today_total // today_count if today_count > 0 else 0
    
    # This week's KPIs (Monday to current day)
    week_start = target_date - __import__('datetime').timedelta(days=target_date.weekday())
    week_start = week_start.replace(hour=0, minute=0, second=0)
    
    week_sales = Sale.query.filter(
        Sale.payment_status == 'PAID',
        Sale.sale_datetime >= week_start,
        Sale.sale_datetime <= today_end
    ).all()
    
    week_total = sum(s.total_tzs for s in week_sales)
    
    # This month's KPIs
    month_start = target_date.replace(day=1, hour=0, minute=0, second=0)
    
    month_sales = Sale.query.filter(
        Sale.payment_status == 'PAID',
        Sale.sale_datetime >= month_start,
        Sale.sale_datetime <= today_end
    ).all()
    
    month_total = sum(s.total_tzs for s in month_sales)
    
    # Inventory health
    all_products = Product.query.filter_by(is_active=True).all()
    low_stock_count = 0
    out_of_stock_count = 0
    
    for product in all_products:
        current_stock = product.get_current_stock()
        if current_stock == 0:
            out_of_stock_count += 1
        elif current_stock < product.reorder_level:
            low_stock_count += 1
    
    return jsonify({
        'status': 'success',
        'date': date_str,
        'today': {
            'total_sales_tzs': today_total,
            'transactions': today_count,
            'avg_basket_tzs': today_avg
        },
        'week': {
            'total_sales_tzs': week_total,
            'transactions': len(week_sales)
        },
        'month': {
            'total_sales_tzs': month_total,
            'transactions': len(month_sales)
        },
        'inventory': {
            'low_stock_count': low_stock_count,
            'out_of_stock_count': out_of_stock_count
        }
    }), 200


@app.route('/api/reports/sales', methods=['GET'])
def api_reports_sales():
    """
    Sales summary report with daily breakdown and payment method breakdown
    
    Query params:
    - from: YYYY-MM-DD
    - to: YYYY-MM-DD
    """
    from_str = request.args.get('from', (datetime.utcnow() - __import__('datetime').timedelta(days=30)).strftime('%Y-%m-%d'))
    to_str = request.args.get('to', datetime.utcnow().strftime('%Y-%m-%d'))
    
    try:
        from_date = datetime.strptime(from_str, '%Y-%m-%d').replace(hour=0, minute=0, second=0)
        to_date = datetime.strptime(to_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    except ValueError:
        return jsonify({'status': 'error', 'message': 'Invalid date format'}), 400
    
    # Get all PAID sales in range
    sales = Sale.query.filter(
        Sale.payment_status == 'PAID',
        Sale.sale_datetime >= from_date,
        Sale.sale_datetime <= to_date
    ).order_by(Sale.sale_datetime).all()
    
    # Daily breakdown
    daily_data = {}
    for sale in sales:
        date_key = sale.sale_datetime.strftime('%Y-%m-%d')
        if date_key not in daily_data:
            daily_data[date_key] = {'total_tzs': 0, 'transactions': 0}
        daily_data[date_key]['total_tzs'] += sale.total_tzs
        daily_data[date_key]['transactions'] += 1
    
    daily_list = [{'date': k, **v} for k, v in sorted(daily_data.items())]
    
    # Payment method breakdown
    payment_breakdown = {}
    for sale in sales:
        method = sale.payment_method
        if method not in payment_breakdown:
            payment_breakdown[method] = {'total_tzs': 0, 'transactions': 0}
        payment_breakdown[method]['total_tzs'] += sale.total_tzs
        payment_breakdown[method]['transactions'] += 1
    
    payment_list = [{'method': k, **v} for k, v in sorted(payment_breakdown.items())]
    
    total_revenue = sum(s.total_tzs for s in sales)
    
    return jsonify({
        'status': 'success',
        'from': from_str,
        'to': to_str,
        'total_revenue_tzs': total_revenue,
        'total_transactions': len(sales),
        'daily': daily_list,
        'payment_methods': payment_list
    }), 200


@app.route('/api/reports/top-products', methods=['GET'])
def api_reports_top_products():
    """
    Top products report - by quantity or revenue
    
    Query params:
    - from: YYYY-MM-DD
    - to: YYYY-MM-DD
    - by: 'qty' or 'revenue' (default: revenue)
    - limit: max results (default 20, max 100)
    """
    from_str = request.args.get('from', (datetime.utcnow() - __import__('datetime').timedelta(days=30)).strftime('%Y-%m-%d'))
    to_str = request.args.get('to', datetime.utcnow().strftime('%Y-%m-%d'))
    by = request.args.get('by', 'revenue').lower()
    limit = min(int(request.args.get('limit', 20)), 100)
    
    try:
        from_date = datetime.strptime(from_str, '%Y-%m-%d').replace(hour=0, minute=0, second=0)
        to_date = datetime.strptime(to_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    except ValueError:
        return jsonify({'status': 'error', 'message': 'Invalid date format'}), 400
    
    # Get all PAID sales in range
    sales = Sale.query.filter(
        Sale.payment_status == 'PAID',
        Sale.sale_datetime >= from_date,
        Sale.sale_datetime <= to_date
    ).all()
    
    # Aggregate by product
    product_data = {}
    for sale in sales:
        for item in sale.items:
            sku = item.sku
            if sku not in product_data:
                product_data[sku] = {
                    'sku': item.sku,
                    'name': item.product_name_snapshot,
                    'qty_sold': 0,
                    'revenue_tzs': 0
                }
            product_data[sku]['qty_sold'] += item.qty
            product_data[sku]['revenue_tzs'] += item.line_total_tzs
    
    # Sort and limit
    if by == 'qty':
        sorted_products = sorted(product_data.values(), key=lambda x: x['qty_sold'], reverse=True)
    else:
        sorted_products = sorted(product_data.values(), key=lambda x: x['revenue_tzs'], reverse=True)
    
    result = sorted_products[:limit]
    
    return jsonify({
        'status': 'success',
        'from': from_str,
        'to': to_str,
        'by': by,
        'count': len(result),
        'products': result
    }), 200


@app.route('/api/reports/inventory-health', methods=['GET'])
def api_reports_inventory_health():
    """
    Inventory health report - low stock and out of stock items
    Also includes inventory value estimate
    
    Returns:
    - low_stock_items: Items below reorder level
    - out_of_stock_items: Items with zero stock
    - inventory_value_estimate_tzs: Sum of (cost_price * current_stock)
    """
    products = Product.query.filter_by(is_active=True).all()
    
    low_stock_items = []
    out_of_stock_items = []
    inventory_value = 0
    
    for product in products:
        current_stock = product.get_current_stock()
        cost = int(float(product.cost_price_tzs))
        item_value = cost * current_stock
        inventory_value += item_value
        
        if current_stock == 0:
            out_of_stock_items.append({
                'sku': product.sku,
                'name': product.name,
                'category': product.category,
                'reorder_level': product.reorder_level
            })
        elif current_stock < product.reorder_level:
            low_stock_items.append({
                'sku': product.sku,
                'name': product.name,
                'category': product.category,
                'current_stock': current_stock,
                'reorder_level': product.reorder_level,
                'units_to_reorder': product.reorder_level - current_stock
            })
    
    return jsonify({
        'status': 'success',
        'low_stock': {
            'count': len(low_stock_items),
            'items': sorted(low_stock_items, key=lambda x: x['units_to_reorder'], reverse=True)
        },
        'out_of_stock': {
            'count': len(out_of_stock_items),
            'items': out_of_stock_items
        },
        'inventory_value_estimate_tzs': inventory_value
    }), 200


# ==================== REPORTS WEB ROUTES ====================

@app.route('/reports/dashboard')
def reports_dashboard():
    """Dashboard page with KPI cards"""
    return render_template('reports/dashboard.html')


@app.route('/reports/sales')
def reports_sales():
    """Sales summary report page"""
    return render_template('reports/sales.html')


@app.route('/reports/top-products')
def reports_top_products():
    """Top products report page"""
    return render_template('reports/top_products.html')


@app.route('/reports/inventory-health')
def reports_inventory_health():
    """Inventory health report page"""
    return render_template('reports/inventory_health.html')


@app.errorhandler(404)
def not_found(error):
    return jsonify({'status': 'error', 'message': 'Not found'}), 404


@app.errorhandler(500)
def server_error(error):
    return jsonify({'status': 'error', 'message': 'Server error'}), 500


# ==================== CLI COMMANDS ====================

@app.cli.command()
def init():
    """Initialize database"""
    init_db()
    print("Database initialized")


@app.cli.command()
def seed():
    """Seed database with sample data"""
    seed_sample_data()
    print("Sample data loaded")



# ==================== AUTO INIT ON DEPLOY (Render/Gunicorn) ====================

def ensure_db_ready():
    """Create tables (and seed demo data if empty) when app starts on Render."""
    with app.app_context():
        db.create_all()
        # Optional: seed demo data only if no products exist
        if Product.query.count() == 0:
            seed_sample_data()

ensure_db_ready()
# ==================== MAIN ====================

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='127.0.0.1', port=5000)
